import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
import re
import os
import webbrowser
import uuid
from mixpanel import Mixpanel

CONFIG_FILE = "config.txt"

# -------------------------------------------------------------------------
# MIXPANEL TELEMETRY CONFIGURATION
# -------------------------------------------------------------------------
# Replace this token with the actual token from your free Mixpanel.com project dashboard
MIXPANEL_TOKEN = "5e1d74ec084e78b8e8abf6cd6502b60d"
mp = Mixpanel(MIXPANEL_TOKEN)

def track_anonymous_usage():
    """Generates a unique hardware ID to track active users and app launches completely anonymously."""
    id_file = "user_id.txt"
    
    # 1. Read or generate a unique random ID for this specific machine
    if os.path.exists(id_file):
        try:
            with open(id_file, "r", encoding="utf-8") as f:
                distinct_id = f.read().strip()
        except Exception:
            distinct_id = str(uuid.uuid4())
    else:
        distinct_id = str(uuid.uuid4())
        try:
            with open(id_file, "w", encoding="utf-8") as f:
                f.write(distinct_id)
        except Exception:
            pass
            
    # 2. Fire the 'App Launched' event silently to your analytics dashboard
    try:
        mp.track(distinct_id, 'App Launched', {
            'Platform': 'Windows/Desktop',
            'App Name': 'Kaabe Dashboard'
        })
    except Exception:
        # Fails silently if the end-user has no active internet connection
        pass


# -------------------------------------------------------------------------
# KAABE WORKSPACE FUNCTIONS
# -------------------------------------------------------------------------
def get_or_set_user_name(force_reset=False):
    """Reads the user's name from config or launches a custom, readable dark-mode input window."""
    if not force_reset and os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                name = f.read().strip()
                if name:
                    return name
        except Exception:
            pass

    dialog = tk.Toplevel() if force_reset else tk.Tk()
    dialog.title("Kaabe - Setup")
    dialog.geometry("500x260")
    dialog.resizable(False, False)
    dialog.config(bg="#121212")

    dialog.update_idletasks()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    x = (dialog.winfo_screenwidth() // 2) - (width // 2)
    y = (dialog.winfo_screenheight() // 2) - (height // 2)
    dialog.geometry(f'{width}x{height}+{x}+{y}')

    user_name_res = ["User"]

    def submit_name(event=None):
        entered = name_entry.get().strip()
        if entered:
            user_name_res[0] = entered
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(user_name_res[0])
        except Exception:
            pass
        dialog.destroy()

    pad_frame = tk.Frame(dialog, bg="#121212", padx=30, pady=25)
    pad_frame.pack(fill="both", expand=True)

    welcome_lbl = tk.Label(
        pad_frame, 
        text="Welcome to Kaabe! ✨", 
        font=("Segoe UI", 16, "bold"), 
        bg="#121212", 
        fg="#ffffff"
    )
    welcome_lbl.pack(anchor="w", pady=(0, 5))

    sub_lbl = tk.Label(
        pad_frame, 
        text="Please enter your name to customize your workspace:", 
        font=("Segoe UI", 11), 
        bg="#121212", 
        fg="#b0b0b0"
    )
    sub_lbl.pack(anchor="w", pady=(0, 20))

    name_entry = tk.Entry(
        pad_frame, 
        font=("Segoe UI", 14), 
        bg="#1e1e1e", 
        fg="#ffffff", 
        insertbackground="#ffffff", 
        bd=1, 
        relief="solid",
        highlightthickness=1,
        highlightbackground="#333333",
        highlightcolor="#2e7d32"
    )
    name_entry.pack(fill="x", ipady=6, pady=(0, 20))
    name_entry.focus_set()

    name_entry.bind("<Return>", submit_name)

    submit_btn = tk.Button(
        pad_frame, 
        text="Get Started  🚀", 
        command=submit_name, 
        font=("Segoe UI", 11, "bold"), 
        bg="#2e7d32", 
        fg="#ffffff", 
        activebackground="#1b5e20", 
        activeforeground="#ffffff", 
        bd=0, 
        cursor="hand2",
        pady=6
    )
    submit_btn.pack(fill="x")

    dialog.grab_set()
    dialog.wait_window()
    return user_name_res[0]


def change_user_name_action():
    """Triggers the profile configuration sequence to overwrite an existing saved name profile."""
    new_name = get_or_set_user_name(force_reset=True)
    if new_name and new_name != "User":
        greeting_label.config(text=f"Hello, {new_name} 👋")


def open_developer_contact():
    """Launches default system web browser directly to your WhatsApp profile chat window."""
    webbrowser.open("https://wa.me/252612008837")


def process_dataframe_core(df_full_scan, output_dir):
    """Core logic to structure, clean, format and save data regardless of source."""
    output_filename = os.path.join(output_dir, "clean_processed_stock.xlsx")
    
    if os.path.exists(output_filename):
        try:
            os.rename(output_filename, output_filename)
        except OSError:
            messagebox.showerror("File Lock Detected", 
                "⚠️ Access Denied! Please close the previous 'clean_processed_stock.xlsx' in Microsoft Excel before running Kaabe again.")
            return False

    date_only_value = datetime.now().strftime("%Y-%m-%d")
    location_value = "Default Location"  
    df_clean = pd.DataFrame()

    matrix_str = df_full_scan.astype(str).to_string()
    is_raw_pasted_mode = "is_raw_pasted_mode" in df_full_scan.columns
    is_lsl_mode = "is_lsl_mode" in df_full_scan.columns
    is_b2c_special_mode = "is_b2c_special_mode" in df_full_scan.columns
    is_trbn_special_mode = "is_trbn_special_mode" in df_full_scan.columns
    
    if is_raw_pasted_mode:
        df_clean['Product'] = df_full_scan['Product']
        df_clean['Closing'] = df_full_scan['Closing']
        if "Detected_Location" in df_full_scan.columns and not df_full_scan["Detected_Location"].dropna().empty:
            location_value = str(df_full_scan["Detected_Location"].iloc[0])
    
    elif "Op/Operation" in matrix_str or "Operation" in matrix_str:
        header_row_idx = None
        for idx, row in df_full_scan.iterrows():
            row_items = [str(x).strip().lower() for x in row.dropna()]
            if any(h in row_items for h in ['product', 'closing', 'op/operation', 'operation', 'quantity', 'on hand']):
                header_row_idx = idx
                break
        
        if header_row_idx is not None:
            new_cols = df_full_scan.iloc[header_row_idx].astype(str).str.strip().tolist()
            df_raw = df_full_scan.iloc[header_row_idx + 1:].copy()
            df_raw.columns = new_cols
        else:
            df_raw = df_full_scan.copy()
            df_raw.columns = [str(c).strip() for c in df_raw.columns]

        prod_col = None
        for col in df_raw.columns:
            if str(col).lower() in ['product', 'product name', 'item']:
                prod_col = col
                break
        if prod_col is None and df_raw.shape[1] > 1:
            prod_col = df_raw.columns[1]

        val_col = None
        for col in df_raw.columns:
            if str(col).lower() in ['closing', 'quantity', 'on hand', 'balance', 'qty']:
                val_col = col
                break
        if val_col is None and df_raw.shape[1] > 3:
            val_col = df_raw.columns[3]

        loc_col = None
        for col in df_raw.columns:
            if any(x in str(col).lower() for x in ['location', 'op/operation', 'operation', 'warehouse']):
                loc_col = col
                break
        
        if loc_col is not None and not df_raw[loc_col].dropna().empty:
            raw_loc = str(df_raw[loc_col].dropna().iloc[0]).strip()
            location_value = raw_loc.split("/")[-1].strip() if "/" in raw_loc else raw_loc

        if prod_col in df_raw.columns and val_col in df_raw.columns:
            df_clean['Product'] = df_raw[prod_col].astype(str).str.strip()
            df_clean['Closing'] = pd.to_numeric(df_raw[val_col], errors='coerce').fillna(0)
        else:
            raise KeyError("Could not dynamically isolate 'Product' or Stock balances from the provided data layout.")

    else:
        header_row_idx = None
        for idx, row in df_full_scan.iterrows():
            row_items = [str(x).strip().lower() for x in row.dropna()]
            if 'product' in row_items or 'closing' in row_items:
                header_row_idx = idx
                break

        for idx, row in df_full_scan.iterrows():
            if header_row_idx is not None and idx >= header_row_idx:
                break
            row_str = " ".join([str(x) for x in row.dropna()])
            if "date:" in row_str.lower() or "date" in row_str.lower():
                cleaned_text = row_str.replace("Date:", "").replace("date:", "").strip()
                date_only_value = cleaned_text.split("TO")[0].strip() if "TO" in cleaned_text else cleaned_text
                break

        if header_row_idx is not None:
            new_cols = df_full_scan.iloc[header_row_idx].astype(str).str.strip().tolist()
            df_raw = df_full_scan.iloc[header_row_idx + 1:].copy()
            df_raw.columns = new_cols
        else:
            df_raw = df_full_scan.copy()
            df_raw.columns = [str(c).strip() for c in df_raw.columns]

        if 'Product' not in df_raw.columns or 'Closing' not in df_raw.columns:
            raise KeyError("Could not find standard columns ('Product' & 'Closing') in this data block.")
            
        location_col = None
        for col in df_raw.columns:
            if 'location' in str(col).lower():
                location_col = col
                break
        
        if location_col is not None and not df_raw[location_col].dropna().empty:
            raw_loc = str(df_raw[location_col].dropna().iloc[0]).strip()
            location_value = raw_loc.split("/")[-1].strip() if "/" in raw_loc else raw_loc

        df_clean['Product'] = df_raw['Product'].astype(str).str.strip()
        df_clean['Closing'] = pd.to_numeric(df_raw['Closing'], errors='coerce').fillna(0)

    df_clean.dropna(subset=['Product'], inplace=True)
    df_clean = df_clean[(df_clean['Product'] != "") & (df_clean['Product'].str.lower() != "nan")]
    
    ignore_keywords = ['total', 'op/operation', 'operation', 'b2c/location', 'trbn/stock', 'inventory report', 'search...', '/']
    for kw in ignore_keywords:
        df_clean = df_clean[~df_clean['Product'].str.lower().str.contains(kw)]
        
    df_clean['Closing'] = pd.to_numeric(df_clean['Closing'], errors='coerce').fillna(0)
    df_clean = df_clean[df_clean['Closing'] != 0]
    
    if df_clean.empty:
        messagebox.showwarning("Empty Results", "Data was processed, but zero items had active, non-zero inventory balances.")
        return False

    df_clean = df_clean.sort_values(by='Product', ascending=True).reset_index(drop=True)
    total_stock = df_clean['Closing'].sum()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Report"
    ws.views.sheetView[0].showGridLines = True

    ws.append([])
    ws.append([])
    
    if is_lsl_mode:
        ws.append(["Location:", location_value, "", ""])
        ws.merge_cells("B3:D3")
        ws.append(["Date:", date_only_value, "", ""])
        ws.merge_cells("B4:D4")
        
        headers = ['Product', 'Closing', 'Counted', 'Sales']
        ws.append(headers)
        
        for index, row in df_clean.iterrows():
            ws.append([row['Product'], row['Closing'], "", ""])
            
        ws.append(['TOTAL :', total_stock, "", ""])
        max_target_cols = 4

    elif is_b2c_special_mode:
        ws.append(["Location:", location_value, "", ""])
        ws.merge_cells("B3:D3")
        ws.append(["Date:", date_only_value, "", ""])
        ws.merge_cells("B4:D4")
        
        headers = ['Product', 'Closing', 'Counted', 'Difference']
        ws.append(headers)
        
        for index, row in df_clean.iterrows():
            curr_row = ws.max_row + 1
            formula_str = f"=C{curr_row}-B{curr_row}"
            ws.append([row['Product'], row['Closing'], row['Closing'], formula_str])
            
        ws.append(['TOTAL :', total_stock, total_stock, "=C{0}-B{0}".format(ws.max_row + 1)])
        max_target_cols = 4

    elif is_trbn_special_mode:
        ws.append(["Location:", location_value, "", ""])
        ws.merge_cells("B3:D3")
        ws.append(["Date:", date_only_value, "", ""])
        ws.merge_cells("B4:D4")
        
        headers = ['Product', 'Closing', 'Counted', 'Difference']
        ws.append(headers)
        
        for index, row in df_clean.iterrows():
            curr_row = ws.max_row + 1
            formula_str = f"=C{curr_row}-B{curr_row}"
            ws.append([row['Product'], row['Closing'], row['Closing'], formula_str])
            
        ws.append(['TOTAL :', total_stock, total_stock, "=C{0}-B{0}".format(ws.max_row + 1)])
        max_target_cols = 4

    else:
        ws.append(["Location:", location_value, ""])
        ws.merge_cells("B3:C3")
        ws.append(["Date:", date_only_value, ""])
        ws.merge_cells("B4:C4")
        
        headers = ['Product', 'Closing', 'Counted']
        ws.append(headers)
        
        for index, row in df_clean.iterrows():
            ws.append([row['Product'], row['Closing'], ""])
            
        ws.append(['TOTAL :', total_stock, ""])
        max_target_cols = 3

    font_regular = Font(name="Calibri", size=12)
    font_bold = Font(name="Calibri", size=12, bold=True)
    crisp_border = Border(
        left=Side(style='medium', color='000000'), 
        right=Side(style='medium', color='000000'), 
        top=Side(style='medium', color='000000'), 
        bottom=Side(style='medium', color='000000')
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_target_cols), start=1):
        for cell in row:
            if row_idx == 3 or row_idx == 4:
                cell.font = font_bold          
                cell.border = crisp_border     
                cell.alignment = Alignment(horizontal='center', vertical='center')
                continue
            if row_idx < 5:
                continue
            if row_idx == 5:
                cell.font = font_bold
                cell.border = crisp_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                continue
            if row_idx == ws.max_row:
                cell.font = font_bold
                cell.border = crisp_border
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                continue

            cell.font = font_regular
            cell.border = crisp_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    max_len_a = max(len(str(cell.value or '')) for cell in ws['A'] if cell.row >= 5)
    ws.column_dimensions['A'].width = max_len_a + 4 

    if is_lsl_mode:
        ws.column_dimensions['B'].width = 12.55  
        ws.column_dimensions['C'].width = 12.55
        ws.column_dimensions['D'].width = 12.55
        ws.print_area = f'A3:D{ws.max_row}'
    elif is_b2c_special_mode:
        ws.column_dimensions['B'].width = 11.00  
        ws.column_dimensions['C'].width = 11.00
        ws.column_dimensions['D'].width = 11.00
        ws.print_area = f'A3:D{ws.max_row}'
    elif is_trbn_special_mode:
        ws.column_dimensions['B'].width = 13.91  
        ws.column_dimensions['C'].width = 13.91
        ws.column_dimensions['D'].width = 13.91
        ws.print_area = f'A3:D{ws.max_row}'
    else:
        ws.column_dimensions['B'].width = 13.22  
        ws.column_dimensions['C'].width = 28.67  
        ws.print_area = f'A3:C{ws.max_row}'

    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.page_margins.top = 0.39
    ws.page_margins.bottom = 0.39
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.header = 0.5
    ws.page_margins.footer = 0.5

    wb.save(output_filename)
    os.startfile(output_filename)
    return True


def reset_file_uploader_ui():
    """Resets the file interface state completely back to default."""
    file_label.config(text="No spreadsheet selected yet.\nClick the upload button below.", font=("Segoe UI", 9, "italic"))
    run_file_btn.config(state=tk.DISABLED)


def run_file_automation(target_file_path):
    try:
        df_full_scan = pd.read_excel(target_file_path, header=None)
        success = process_dataframe_core(df_full_scan, os.path.dirname(target_file_path))
        if success:
            reset_file_uploader_ui()
    except Exception as e:
        messagebox.showerror("Automation Error", f"Kaabe halted safely:\n\n{str(e)}")


def run_text_automation(event=None):
    raw_content = text_box.get("1.0", tk.END).strip()
    if not raw_content:
        if event is None:
            messagebox.showwarning("Empty Paste Field", "Please paste your copied Odoo text into the box first!")
        return "break"
    
    try:
        lines = [line.strip() for line in raw_content.split('\n') if line.strip()]
        
        is_b2c_special_environment = False
        is_trbn_special_environment = False
        
        if lines and lines[0].lower() == "b2c2":
            is_b2c_special_environment = True
            lines = lines[1:]  
        elif lines and lines[0].lower() == "tr2":
            is_trbn_special_environment = True
            lines = lines[1:] 
            
        is_lsl_environment = False
        if not is_b2c_special_environment and not is_trbn_special_environment:
            is_lsl_environment = "lsl" in raw_content.lower() or "local sales" in raw_content.lower()
        
        products = []
        closings = []
        
        if is_b2c_special_environment:
            detected_location = "B2C Store"
        elif is_trbn_special_environment:
            detected_location = "TRBN Store"
        elif is_lsl_environment:
            detected_location = "Local Sales"
        else:
            detected_location = "Default Location"
        
        for line in lines:
            if is_lsl_environment:
                if any(k in line.lower() for k in ["inventory report", "search...", "1-75", "1-49", "1-80", "1-93"]):
                    continue
                
                line_clean = re.sub(r'(?i)\s*view\s+details\s*$', '', line).strip()
                parts = line_clean.split('\t') if '\t' in line_clean else line_clean.split()
                if len(parts) >= 2:
                    try:
                        qty_str = parts[-1].replace(",", "")
                        qty_val = float(qty_str)
                        
                        if parts[0].lower() in ["lsl", "lsl/local", "sales", "location"] or "/" in parts[0]:
                            remaining_text = " ".join(parts[1:-1])
                            if "local sales location" in remaining_text.lower():
                                remaining_text = re.sub(r'(?i)^local\s+sales\s+location\s*', '', remaining_text).strip()
                        else:
                            remaining_text = " ".join(parts[:-1])
                            
                        clean_prod_name = remaining_text.strip()
                        if clean_prod_name:
                            products.append(clean_prod_name)
                            closings.append(qty_val)
                    except ValueError:
                        pass
                continue

            if is_b2c_special_environment:
                if any(k in line.lower() for k in ["inventory report", "search...", "1-75", "1-49", "1-80", "1-93"]):
                    continue
                
                parts = line.split('\t') if '\t' in line else line.split()
                if len(parts) >= 2:
                    try:
                        qty_str = parts[-1].replace(",", "")
                        qty_val = float(qty_str)
                        
                        if "/" in parts[0] or parts[0].lower() in ["b2c", "b2c/stock", "b2c/location"]:
                            remaining_text = " ".join(parts[1:-1])
                        else:
                            remaining_text = " ".join(parts[:-1])
                            
                        clean_prod_name = remaining_text.strip()
                        if clean_prod_name:
                            products.append(clean_prod_name)
                            closings.append(qty_val)
                    except ValueError:
                        pass
                continue

            if is_trbn_special_environment:
                if any(k in line.lower() for k in ["inventory report", "search...", "1-75", "1-49", "1-80", "1-93"]):
                    continue
                
                parts = line.split('\t') if '\t' in line else line.split()
                if len(parts) >= 2:
                    try:
                        qty_str = parts[-1].replace(",", "")
                        qty_val = float(qty_str)
                        
                        if "/" in parts[0] or parts[0].lower() in ["trbn", "trbn/stock", "trbn/location"]:
                            remaining_text = " ".join(parts[1:-1])
                        else:
                            remaining_text = " ".join(parts[:-1])
                            
                        clean_prod_name = remaining_text.strip()
                        if clean_prod_name:
                            products.append(clean_prod_name)
                            closings.append(qty_val)
                    except ValueError:
                        pass
                continue

            if any(k in line.lower() for k in ["inventory report", "search...", "1-75", "1-49", "1-80", "1-93"]):
                if len(line.split()) == 1 and line.isalnum():
                    detected_location = line.upper()
                continue
            
            match = re.match(r"^([^\s]+)\s+(.*?)\s+([\d,.]+)\s*$", line)
            if match:
                prefix = match.group(1)
                prod_name = match.group(2).strip()
                qty_str = match.group(3).replace(",", "")
                
                if "/" in prefix:
                    detected_location = prefix.split("/")[0].strip()
                else:
                    detected_location = prefix
                
                products.append(prod_name)
                closings.append(qty_str)
            else:
                parts = line.split()
                if len(parts) >= 2:
                    try:
                        qty = float(parts[-1].replace(",", ""))
                        prod_name = " ".join(parts[:-1])
                        products.append(prod_name)
                        closings.append(qty)
                    except ValueError:
                        pass

        if not products:
            raise KeyError("Could not find standard columns ('Product' & 'Closing') in this data block.")

        df_text = pd.DataFrame({
            'Product': products,
            'Closing': closings,
            'is_raw_pasted_mode': True,
            'Detected_Location': detected_location
        })
        
        if is_lsl_environment:
            df_text['is_lsl_mode'] = True
        elif is_b2c_special_environment:
            df_text['is_b2c_special_mode'] = True
        elif is_trbn_special_environment:
            df_text['is_trbn_special_mode'] = True
        
        default_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
        if not os.path.exists(default_dir):
            default_dir = os.path.expanduser('~')
            
        success = process_dataframe_core(df_text, default_dir)
        if success:
            text_box.delete("1.0", tk.END)
            return "break"
            
    except Exception as e:
        messagebox.showerror("Text Parsing Error", f"Kaabe couldn't map that plain-text data block:\n\n{str(e)}")
    return "break"


def apply_theme_styles():
    if theme_var.get():
        bg_main = "#0d1117"       
        bg_card = "#161b22"       
        bg_input = "#090d13"      
        text_primary = "#c9d1d9"  
        text_muted = "#8b949e"    
        accent_bg = "#238636"     
        accent_hover = "#2ea043"  
        
        root.config(bg=bg_main)
        main_frame.configure(style="Custom.TFrame")
        header_frame.configure(style="Custom.TFrame")
        cards_container.configure(style="Custom.TFrame")
        right_actions_frame.config(bg=bg_main)
        
        file_card.config(bg=bg_card, highlightbackground="#30363d")
        file_inner.config(bg=bg_card)
        text_card.config(bg=bg_card, highlightbackground="#30363d")
        text_inner.config(bg=bg_card)
        
        style.configure("Custom.TFrame", background=bg_main)
        style.configure("TLabel", background=bg_main, foreground=text_primary)
        style.configure("Greeting.TLabel", background=bg_main, foreground="#58a6ff", font=("Segoe UI", 10, "italic")) 
        style.configure("Title.TLabel", background=bg_main, foreground="#ffffff", font=("Segoe UI", 18, "bold"))
        
        lbl1.config(bg=bg_card, fg="#ffffff")
        lbl2.config(bg=bg_card, fg=text_muted)
        lbl3.config(bg=bg_card, fg="#ffffff")
        lbl4.config(bg=bg_card, fg=text_muted)
        
        style.configure("Theme.TCheckbutton", background=bg_main, foreground=text_primary, font=("Segoe UI", 10, "bold"))
        theme_switch.config(text="☀️ Light UI")
        
        file_label.config(bg=bg_input, fg=text_primary, highlightbackground="#30363d")
        text_box.config(bg=bg_input, fg="#f0f6fc", insertbackground="#ffffff", highlightbackground="#30363d")
        
        style.configure("Browse.TButton", background="#21262d", foreground="#c9d1d9", font=("Segoe UI", 10, "bold"))
        style.map("Browse.TButton", background=[('active', '#30363d')])
        
        whatsapp_btn.config(bg="#21262d", fg="#25d366", activebackground="#30363d", activeforeground="#25d366")
        
        style.configure("Run.TButton", background=accent_bg, foreground="#ffffff", font=("Segoe UI", 10, "bold"))
        style.map("Run.TButton", background=[('disabled', '#21262d'), ('active', accent_hover)], foreground=[('disabled', '#484f58')])
        
        reset_lbl.config(bg=bg_main, fg="#58a6ff", activebackground=bg_main, activeforeground="#2ea043")
    else:
        bg_main = "#f6f8fa"       
        bg_card = "#ffffff"       
        bg_input = "#f3f4f6"      
        text_primary = "#24292f"  
        text_muted = "#57606a"    
        accent_bg = "#1a7f37"     
        accent_hover = "#116629"  
        
        root.config(bg=bg_main)
        main_frame.configure(style="Custom.TFrame")
        header_frame.configure(style="Custom.TFrame")
        cards_container.configure(style="Custom.TFrame")
        right_actions_frame.config(bg=bg_main)
        
        file_card.config(bg=bg_card, highlightbackground="#d0d7de")
        file_inner.config(bg=bg_card)
        text_card.config(bg=bg_card, highlightbackground="#d0d7de")
        text_inner.config(bg=bg_card)
        
        style.configure("Custom.TFrame", background=bg_main)
        style.configure("TLabel", background=bg_main, foreground=text_primary)
        style.configure("Greeting.TLabel", background=bg_main, foreground="#0969da", font=("Segoe UI", 10, "italic"))
        style.configure("Title.TLabel", background=bg_main, foreground="#1f2328", font=("Segoe UI", 18, "bold"))
        
        lbl1.config(bg=bg_card, fg="#1f2328")
        lbl2.config(bg=bg_card, fg=text_muted)
        lbl3.config(bg=bg_card, fg="#1f2328")
        lbl4.config(bg=bg_card, fg=text_muted)
        
        style.configure("Theme.TCheckbutton", background=bg_main, foreground=text_primary, font=("Segoe UI", 10, "bold"))
        theme_switch.config(text="🌙 Dark UI")
        
        file_label.config(bg=bg_input, fg=text_primary, highlightbackground="#d0d7de")
        text_box.config(bg="#ffffff", fg=text_primary, insertbackground="#000000", highlightbackground="#d0d7de")
        
        style.configure("Browse.TButton", background="#f3f4f6", foreground=text_primary, font=("Segoe UI", 10, "bold"))
        style.map("Browse.TButton", background=[('active', '#ebecf0')])
        
        whatsapp_btn.config(bg="#f3f4f6", fg="#128c7e", activebackground="#ebecf0", activeforeground="#128c7e")
        
        style.configure("Run.TButton", background=accent_bg, foreground="#ffffff", font=("Segoe UI", 10, "bold"))
        style.map("Run.TButton", background=[('disabled', '#eaeef2'), ('active', accent_hover)], foreground=[('disabled', '#8c959f')])
        
        reset_lbl.config(bg=bg_main, fg="#0969da", activebackground=bg_main, activeforeground="#116629")

def browse_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if filepath:
        file_label.config(text=f"📂 File Selected:\n{os.path.basename(filepath)}", font=("Segoe UI", 9, "bold"))
        run_file_btn.config(state=tk.NORMAL)
        run_file_btn.config(command=lambda: run_file_automation(filepath))


ACTIVE_USER_NAME = get_or_set_user_name()

root = tk.Tk()
root.title("Kaabe Dashboard")
root.geometry("860x490") 
root.resizable(False, False)

# Initialize global tracking variables before constructing widgets
theme_var = tk.BooleanVar(value=True) 

style = ttk.Style(root)
style.theme_use("clam")  

main_frame = ttk.Frame(root, padding=20)
main_frame.pack(fill="both", expand=True)

# Top Bar Header
header_frame = ttk.Frame(main_frame)
header_frame.pack(fill="x", pady=(0, 15))

title_text_frame = ttk.Frame(header_frame, style="Custom.TFrame")
title_text_frame.pack(side="left", anchor="w")

greeting_container = tk.Frame(title_text_frame, bg="#0d1117") 
greeting_container.pack(anchor="w", pady=(0, 1))

greeting_label = ttk.Label(greeting_container, text=f"Hello, {ACTIVE_USER_NAME} 👋", style="Greeting.TLabel")
greeting_label.pack(side="left")

reset_lbl = tk.Button(
    greeting_container, 
    text="(Edit Name)", 
    font=("Segoe UI", 9, "underline"), 
    bd=0, 
    cursor="hand2", 
    command=change_user_name_action
)
reset_lbl.pack(side="left", padx=(8, 0))

title_label = ttk.Label(title_text_frame, text="Kaabe Inventory Workspace", style="Title.TLabel")
title_label.pack(anchor="w")

# Right Actions Header Frame
right_actions_frame = tk.Frame(header_frame, bg="#0d1117")
right_actions_frame.pack(side="right", anchor="e", padx=5)

theme_switch = ttk.Checkbutton(right_actions_frame, style="Theme.TCheckbutton", variable=theme_var, command=apply_theme_styles)
theme_switch.pack(side="left", padx=(0, 12))

# WhatsApp Action Icon
whatsapp_btn = tk.Button(
    right_actions_frame, 
    text="💬 WhatsApp", 
    command=open_developer_contact, 
    font=("Segoe UI", 10, "bold"),
    bd=1,
    relief="solid",
    padx=10,
    pady=3,
    cursor="hand2",
    highlightthickness=0
)
whatsapp_btn.pack(side="left")

# Side-by-Side Application Cards Layout
cards_container = ttk.Frame(main_frame)
cards_container.pack(fill="both", expand=True)

# --- LEFT CARD: FILE UPLOADER ---
file_card = tk.LabelFrame(cards_container, bd=0, relief="solid", highlightthickness=1)
file_card.pack(side="left", fill="both", expand=True, padx=(0, 10), pady=5)

file_inner = tk.Frame(file_card, bd=0)
file_inner.pack(fill="both", expand=True, padx=20, pady=20)

lbl1 = tk.Label(file_inner, text="📁 Local Spreadsheet Import", font=("Segoe UI", 12, "bold"))
lbl1.pack(anchor="w", pady=(0, 2))
lbl2 = tk.Label(file_inner, text="Drop standard Odoo Excel exports here.", font=("Segoe UI", 9))
lbl2.pack(anchor="w", pady=(0, 20))

file_label = tk.Label(file_inner, text="No spreadsheet selected yet.\nClick the upload button below.", font=("Segoe UI", 9, "italic"), height=4, anchor="center", justify="center", bd=1, relief="solid", highlightthickness=1)
file_label.pack(fill="x", pady=(0, 25))

browse_btn = ttk.Button(file_inner, text="📁 Select Excel File", command=browse_file, style="Browse.TButton")
browse_btn.pack(fill="x", ipady=5, pady=(0, 10))

run_file_btn = ttk.Button(file_inner, text="🚀 Process Document", state=tk.DISABLED, style="Run.TButton")
run_file_btn.pack(fill="x", ipady=5)


# --- RIGHT CARD: RAW TEXT COPIER ---
text_card = tk.LabelFrame(cards_container, bd=0, relief="solid", highlightthickness=1)
text_card.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=5)

text_inner = tk.Frame(text_card, bd=0)
text_inner.pack(fill="both", expand=True, padx=20, pady=20)

lbl3 = tk.Label(text_inner, text="📝 Instant Paste Console", font=("Segoe UI", 12, "bold"))
lbl3.pack(anchor="w", pady=(0, 2))
lbl4 = tk.Label(text_inner, text="Paste raw grid text directly from your web screen browser.", font=("Segoe UI", 9))
lbl4.pack(anchor="w", pady=(0, 10))

text_box = tk.Text(text_inner, font=("Consolas", 10), height=8, bd=1, relief="solid", highlightthickness=1, wrap="none")
text_box.pack(fill="both", expand=True, pady=(0, 15))

text_box.bind("<Return>", run_text_automation)

run_text_btn = ttk.Button(text_inner, text="⚡ Process Pasted Data", command=run_text_automation, style="Run.TButton")
run_text_btn.pack(fill="x", ipady=5)

apply_theme_styles()

# Trigger active metrics ping right before window visualization loops
track_anonymous_usage()

root.mainloop()